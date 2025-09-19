import os
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import json

class HealthcareQuestionsExporter:
    """
    A class to export healthcare questions data to various formats
    with advanced formatting and visualization options.
    """
    
    def __init__(self, data_dir="healthcare_data/data"):
        """Initialize the exporter with data directory"""
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)
    
    def load_csv(self, csv_path):
        """Load data from a CSV file"""
        try:
            df = pd.read_csv(csv_path)
            print(f"Loaded {len(df)} questions from {csv_path}")
            return df
        except Exception as e:
            print(f"Error loading CSV file: {e}")
            return None
    
    def export_to_excel(self, df, output_filename="healthcare_questions.xlsx"):
        """
        Export data to Excel with color coding and formatting
        """
        output_path = os.path.join(self.data_dir, output_filename)
        
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Healthcare Questions"
        
        # Add data to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply color coding to rows based on healthcare professional type
                elif c_idx == df.columns.get_loc("Color") + 1 and r_idx > 1:
                    color_code = value.lstrip('#')
                    
                    # Apply the color to the entire row
                    for col in range(1, len(row) + 1):
                        if col != c_idx:  # Skip coloring the color code cell itself
                            ws.cell(row=r_idx, column=col).fill = PatternFill(
                                start_color=color_code, 
                                end_color=color_code, 
                                fill_type="solid"
                            )
                            
                            # Make text white if background is dark
                            if self._is_dark_color(color_code):
                                ws.cell(row=r_idx, column=col).font = Font(color="FFFFFF")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Save the workbook
        wb.save(output_path)
        print(f"Excel file created at {output_path}")
        return output_path
    
    def _is_dark_color(self, hex_color):
        """
        Determine if a color is dark (to decide text color)
        """
        # Convert hex to RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        
        # Calculate luminance
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        
        # Return True if color is dark
        return luminance < 0.5
    
    def create_visualizations(self, df, output_dir=None):
        """
        Create visualizations of the healthcare questions data
        """
        if output_dir is None:
            output_dir = self.data_dir
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Set the style
        sns.set(style="whitegrid")
        
        # 1. Healthcare Professional Type Distribution
        plt.figure(figsize=(12, 8))
        type_counts = df['Healthcare Professional Type'].value_counts()
        ax = sns.barplot(x=type_counts.index, y=type_counts.values)
        plt.title('Questions by Healthcare Professional Type', fontsize=16)
        plt.xlabel('Healthcare Professional Type', fontsize=12)
        plt.ylabel('Number of Questions', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        
        # Add count labels on top of bars
        for i, count in enumerate(type_counts.values):
            ax.text(i, count + 0.5, str(count), ha='center', fontsize=10)
        
        plt.savefig(os.path.join(output_dir, 'healthcare_type_distribution.png'), dpi=300)
        plt.close()
        
        # 2. Visit Type Distribution (top 10)
        plt.figure(figsize=(12, 8))
        visit_counts = df['Visit Type'].value_counts().head(10)
        ax = sns.barplot(x=visit_counts.index, y=visit_counts.values)
        plt.title('Top 10 Visit Types', fontsize=16)
        plt.xlabel('Visit Type', fontsize=12)
        plt.ylabel('Number of Questions', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        
        # Add count labels on top of bars
        for i, count in enumerate(visit_counts.values):
            ax.text(i, count + 0.5, str(count), ha='center', fontsize=10)
        
        plt.savefig(os.path.join(output_dir, 'visit_type_distribution.png'), dpi=300)
        plt.close()
        
        # 3. Target Audience Distribution
        plt.figure(figsize=(10, 6))
        audience_counts = df['Target Audience'].value_counts()
        ax = sns.barplot(x=audience_counts.index, y=audience_counts.values)
        plt.title('Questions by Target Audience', fontsize=16)
        plt.xlabel('Target Audience', fontsize=12)
        plt.ylabel('Number of Questions', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        
        # Add count labels on top of bars
        for i, count in enumerate(audience_counts.values):
            ax.text(i, count + 0.5, str(count), ha='center', fontsize=10)
        
        plt.savefig(os.path.join(output_dir, 'audience_distribution.png'), dpi=300)
        plt.close()
        
        # 4. Category Distribution (if available)
        if 'Category' in df.columns:
            plt.figure(figsize=(10, 6))
            category_counts = df['Category'].value_counts()
            ax = sns.barplot(x=category_counts.index, y=category_counts.values)
            plt.title('Questions by Category', fontsize=16)
            plt.xlabel('Category', fontsize=12)
            plt.ylabel('Number of Questions', fontsize=12)
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            
            # Add count labels on top of bars
            for i, count in enumerate(category_counts.values):
                ax.text(i, count + 0.5, str(count), ha='center', fontsize=10)
            
            plt.savefig(os.path.join(output_dir, 'category_distribution.png'), dpi=300)
            plt.close()
        
        # 5. Heatmap of Healthcare Type vs Visit Type (top combinations)
        plt.figure(figsize=(14, 10))
        heatmap_data = pd.crosstab(df['Healthcare Professional Type'], df['Visit Type'])
        
        # Select top 10 healthcare types and visit types by frequency
        top_healthcare_types = df['Healthcare Professional Type'].value_counts().head(10).index
        top_visit_types = df['Visit Type'].value_counts().head(10).index
        
        # Filter heatmap data
        heatmap_filtered = heatmap_data.loc[
            heatmap_data.index.isin(top_healthcare_types),
            heatmap_data.columns.isin(top_visit_types)
        ]
        
        sns.heatmap(heatmap_filtered, annot=True, cmap="YlGnBu", fmt='d')
        plt.title('Heatmap: Healthcare Professional Type vs Visit Type', fontsize=16)
        plt.xlabel('Visit Type', fontsize=12)
        plt.ylabel('Healthcare Professional Type', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'heatmap_type_vs_visit.png'), dpi=300)
        plt.close()
        
        print(f"Visualizations created in {output_dir}")
        return output_dir
    
    def export_to_html(self, df, output_filename="healthcare_questions.html"):
        """
        Export data to an interactive HTML table with color coding
        """
        output_path = os.path.join(self.data_dir, output_filename)
        
        # Create a styled HTML table
        html_content = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Healthcare Questions Database</title>
            <link rel="stylesheet" href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css">
            <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
            <script src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    padding: 0;
                    background-color: #f5f5f5;
                }
                .container {
                    max-width: 1200px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 20px;
                    border-radius: 5px;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }
                h1 {
                    color: #333;
                    text-align: center;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                }
                th {
                    background-color: #4F81BD;
                    color: white;
                    padding: 10px;
                    text-align: left;
                }
                td {
                    padding: 8px;
                    border-bottom: 1px solid #ddd;
                }
                tr:hover {
                    background-color: #f5f5f5;
                }
                .filters {
                    margin: 20px 0;
                    display: flex;
                    flex-wrap: wrap;
                    gap: 10px;
                }
                .filter-group {
                    flex: 1;
                    min-width: 200px;
                }
                select, input {
                    width: 100%;
                    padding: 8px;
                    border: 1px solid #ddd;
                    border-radius: 4px;
                }
                label {
                    display: block;
                    margin-bottom: 5px;
                    font-weight: bold;
                }
                .stats {
                    margin: 20px 0;
                    padding: 15px;
                    background-color: #e9f7fe;
                    border-radius: 5px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Healthcare Questions Database</h1>
                
                <div class="stats">
                    <h3>Database Statistics</h3>
                    <p><strong>Total Questions:</strong> """ + str(len(df)) + """</p>
                    <p><strong>Healthcare Professional Types:</strong> """ + str(df['Healthcare Professional Type'].nunique()) + """</p>
                    <p><strong>Visit Types:</strong> """ + str(df['Visit Type'].nunique()) + """</p>
                </div>
                
                <div class="filters">
                    <div class="filter-group">
                        <label for="healthcareTypeFilter">Filter by Healthcare Professional Type:</label>
                        <select id="healthcareTypeFilter">
                            <option value="">All Types</option>
                            """ + ''.join([f'<option value="{type}">{type}</option>' for type in sorted(df['Healthcare Professional Type'].unique())]) + """
                        </select>
                    </div>
                    <div class="filter-group">
                        <label for="visitTypeFilter">Filter by Visit Type:</label>
                        <select id="visitTypeFilter">
                            <option value="">All Visit Types</option>
                            """ + ''.join([f'<option value="{type}">{type}</option>' for type in sorted(df['Visit Type'].unique())]) + """
                        </select>
                    </div>
                    <div class="filter-group">
                        <label for="searchFilter">Search Questions:</label>
                        <input type="text" id="searchFilter" placeholder="Type to search...">
                    </div>
                </div>
                
                <table id="questionsTable" class="display">
                    <thead>
                        <tr>
                            <th>Question</th>
                            <th>Healthcare Professional Type</th>
                            <th>Visit Type</th>
                            <th>Target Audience</th>
                            <th>Source</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        # Add rows with color coding
        for _, row in df.iterrows():
            color = row['Color']
            text_color = '#FFFFFF' if self._is_dark_color(color.lstrip('#')) else '#000000'
            
            html_content += f"""
                <tr style="background-color: {color}; color: {text_color};">
                    <td>{row['Question']}</td>
                    <td>{row['Healthcare Professional Type']}</td>
                    <td>{row['Visit Type']}</td>
                    <td>{row['Target Audience']}</td>
                    <td>{row['Source']}</td>
                </tr>
            """
        
        # Close the HTML structure and add JavaScript for filtering
        html_content += """
                    </tbody>
                </table>
            </div>
            
            <script>
                $(document).ready(function() {
                    // Initialize DataTable
                    var table = $('#questionsTable').DataTable({
                        pageLength: 25,
                        lengthMenu: [[10, 25, 50, 100, -1], [10, 25, 50, 100, "All"]]
                    });
                    
                    // Apply filters
                    $('#healthcareTypeFilter').on('change', function() {
                        table.column(1).search(this.value).draw();
                    });
                    
                    $('#visitTypeFilter').on('change', function() {
                        table.column(2).search(this.value).draw();
                    });
                    
                    $('#searchFilter').on('keyup', function() {
                        table.search(this.value).draw();
                    });
                });
            </script>
        </body>
        </html>
        """
        
        # Write the HTML file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"HTML file created at {output_path}")
        return output_path
    
    def export_to_json(self, df, output_filename="healthcare_questions.json"):
        """
        Export data to a JSON file
        """
        output_path = os.path.join(self.data_dir, output_filename)
        
        # Convert DataFrame to JSON
        json_data = df.to_dict(orient='records')
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2)
        
        print(f"JSON file created at {output_path}")
        return output_path
    
    def create_dashboard(self, df, output_filename="dashboard.html"):
        """
        Create an HTML dashboard with visualizations and interactive table
        """
        # First create the visualizations
        viz_dir = os.path.join(self.data_dir, "visualizations")
        self.create_visualizations(df, viz_dir)
        
        # Get relative paths to visualization images
        viz_paths = {
            'healthcare_type': "visualizations/healthcare_type_distribution.png",
            'visit_type': "visualizations/visit_type_distribution.png",
            'audience': "visualizations/audience_distribution.png"
        }
        
        if 'Category' in df.columns:
            viz_paths['category'] = "visualizations/category_distribution.png"
        
        viz_paths['heatmap'] = "visualizations/heatmap_type_vs_visit.png"
        
        # Create the dashboard HTML
        output_path = os.path.join(self.data_dir, output_filename)
        
        dashboard_html = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Healthcare Questions Dashboard</title>
            <link rel="stylesheet" href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css">
            <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
            <script src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f5f5f5;
                }
                .container {
                    max-width: 1200px;
                    margin: 0 auto;
                    padding: 20px;
                }
                .dashboard-header {
                    background-color: #4F81BD;
                    color: white;
                    padding: 20px;
                    text-align: center;
                    margin-bottom: 20px;
                }
                .stats-panel {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 20px;
                    margin-bottom: 30px;
                }
                .stat-card {
                    flex: 1;
                    min-width: 200px;
                    background-color: white;
                    border-radius: 5px;
                    padding: 20px;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                    text-align: center;
                }
                .stat-value {
                    font-size: 2em;
                    font-weight: bold;
                    color: #4F81BD;
                }
                .stat-label {
                    color: #666;
                    margin-top: 10px;
                }
                .viz-panel {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 20px;
                    margin-bottom: 30px;
                }
                .viz-card {
                    flex: 1 1 calc(50% - 20px);
                    min-width: 300px;
                    background-color: white;
                    border-radius: 5px;
                    padding: 20px;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }
                .viz-title {
                    font-size: 1.2em;
                    margin-bottom: 15px;
                    color: #333;
                }
                .viz-img {
                    width: 100%;
                    height: auto;
                }
                .data-table-section {
                    background-color: white;
                    border-radius: 5px;
                    padding: 20px;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }
                h1, h2, h3 {
                    color: #333;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                }
                th {
                    background-color: #4F81BD;
                    color: white;
                    padding: 10px;
                    text-align: left;
                }
                td {
                    padding: 8px;
                    border-bottom: 1px solid #ddd;
                }
                .filters {
                    margin: 20px 0;
                    display: flex;
                    flex-wrap: wrap;
                    gap: 10px;
                }
                .filter-group {
                    flex: 1;
                    min-width: 200px;
                }
                select, input {
                    width: 100%;
                    padding: 8px;
                    border: 1px solid #ddd;
                    border-radius: 4px;
                }
                label {
                    display: block;
                    margin-bottom: 5px;
                    font-weight: bold;
                }
            </style>
        </head>
        <body>
            <div class="dashboard-header">
                <h1>Healthcare Questions Dashboard</h1>
                <p>Interactive dashboard for exploring healthcare questions by professional type and visit type</p>
            </div>
            
            <div class="container">
                <div class="stats-panel">
                    <div class="stat-card">
                        <div class="stat-value">""" + str(len(df)) + """</div>
                        <div class="stat-label">Total Questions</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + str(df['Healthcare Professional Type'].nunique()) + """</div>
                        <div class="stat-label">Healthcare Professional Types</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + str(df['Visit Type'].nunique()) + """</div>
                        <div class="stat-label">Visit Types</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + str(df['Target Audience'].nunique()) + """</div>
                        <div class="stat-label">Target Audiences</div>
                    </div>
                </div>
                
                <div class="viz-panel">
                    <div class="viz-card">
                        <div class="viz-title">Questions by Healthcare Professional Type</div>
                        <img class="viz-img" src=\"""" + viz_paths['healthcare_type'] + """\" alt="Healthcare Type Distribution">
                    </div>
                    <div class="viz-card">
                        <div class="viz-title">Top 10 Visit Types</div>
                        <img class="viz-img" src=\"""" + viz_paths['visit_type'] + """\" alt="Visit Type Distribution">
                    </div>
                </div>
                
                <div class="viz-panel">
                    <div class="viz-card">
                        <div class="viz-title">Questions by Target Audience</div>
                        <img class="viz-img" src=\"""" + viz_paths['audience'] + """\" alt="Target Audience Distribution">
                    </div>
                    <div class="viz-card">
                        <div class="viz-title">Healthcare Type vs Visit Type</div>
                        <img class="viz-img" src=\"""" + viz_paths['heatmap'] + """\" alt="Heatmap">
                    </div>
                </div>
        """
        
        # Add category visualization if available
        if 'Category' in df.columns:
            dashboard_html += """
                <div class="viz-panel">
                    <div class="viz-card">
                        <div class="viz-title">Questions by Category</div>
                        <img class="viz-img" src=\"""" + viz_paths['category'] + """\" alt="Category Distribution">
                    </div>
                </div>
            """
        
        # Add data table section
        dashboard_html += """
                <div class="data-table-section">
                    <h2>Healthcare Questions Database</h2>
                    
                    <div class="filters">
                        <div class="filter-group">
                            <label for="healthcareTypeFilter">Filter by Healthcare Professional Type:</label>
                            <select id="healthcareTypeFilter">
                                <option value="">All Types</option>
                                """ + ''.join([f'<option value="{type}">{type}</option>' for type in sorted(df['Healthcare Professional Type'].unique())]) + """
                            </select>
                        </div>
                        <div class="filter-group">
                            <label for="visitTypeFilter">Filter by Visit Type:</label>
                            <select id="visitTypeFilter">
                                <option value="">All Visit Types</option>
                                """ + ''.join([f'<option value="{type}">{type}</option>' for type in sorted(df['Visit Type'].unique())]) + """
                            </select>
                        </div>
                        <div class="filter-group">
                            <label for="searchFilter">Search Questions:</label>
                            <input type="text" id="searchFilter" placeholder="Type to search...">
                        </div>
                    </div>
                    
                    <table id="questionsTable" class="display">
                        <thead>
                            <tr>
                                <th>Question</th>
                                <th>Healthcare Professional Type</th>
                                <th>Visit Type</th>
                                <th>Target Audience</th>
                                <th>Source</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        # Add rows with color coding
        for _, row in df.iterrows():
            color = row['Color']
            text_color = '#FFFFFF' if self._is_dark_color(color.lstrip('#')) else '#000000'
            
            dashboard_html += f"""
                            <tr style="background-color: {color}; color: {text_color};">
                                <td>{row['Question']}</td>
                                <td>{row['Healthcare Professional Type']}</td>
                                <td>{row['Visit Type']}</td>
                                <td>{row['Target Audience']}</td>
                                <td>{row['Source']}</td>
                            </tr>
            """
        
        # Close the HTML structure and add JavaScript for filtering
        dashboard_html += """
                        </tbody>
                    </table>
                </div>
            </div>
            
            <script>
                $(document).ready(function() {
                    // Initialize DataTable
                    var table = $('#questionsTable').DataTable({
                        pageLength: 10,
                        lengthMenu: [[10, 25, 50, 100, -1], [10, 25, 50, 100, "All"]]
                    });
                    
                    // Apply filters
                    $('#healthcareTypeFilter').on('change', function() {
                        table.column(1).search(this.value).draw();
                    });
                    
                    $('#visitTypeFilter').on('change', function() {
                        table.column(2).search(this.value).draw();
                    });
                    
                    $('#searchFilter').on('keyup', function() {
                        table.search(this.value).draw();
                    });
                });
            </script>
        </body>
        </html>
        """
        
        # Write the HTML file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(dashboard_html)
        
        print(f"Dashboard created at {output_path}")
        return output_path


# Example usage
def main():
    # Initialize the exporter
    exporter = HealthcareQuestionsExporter()
    
    # Load data from CSV
    csv_path = "healthcare_data/data/healthcare_questions.csv"
    df = exporter.load_csv(csv_path)
    
    if df is not None:
        # Export to Excel with color coding
        excel_path = exporter.export_to_excel(df)
        
        # Create visualizations
        viz_dir = exporter.create_visualizations(df)
        
        # Export to HTML
        html_path = exporter.export_to_html(df)
        
        # Export to JSON
        json_path = exporter.export_to_json(df)
        
        # Create dashboard
        dashboard_path = exporter.create_dashboard(df)
        
        print("\nExport Summary:")
        print(f"Excel file: {excel_path}")
        print(f"Visualizations: {viz_dir}")
        print(f"HTML file: {html_path}")
        print(f"JSON file: {json_path}")
        print(f"Dashboard: {dashboard_path}")


if __name__ == "__main__":
    main()
